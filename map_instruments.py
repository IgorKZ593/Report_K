#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
map_instruments.py — Этап 1.
Каркас: поиск входного JSON isin_Фамилия И.О._start__end.json в Data_work,
чтение и валидация, печать статуса. Никакой логики справочников/записи на диск.
"""

import os
import sys
import json
import re
import shutil
from datetime import datetime
from glob import glob
from pathlib import Path
from typing import Tuple, List, Dict, Any, Optional

# === Автоустановка rich для цветного вывода ===
try:
    from rich.console import Console
    from rich import print
    from rich.table import Table
except ImportError:
    os.system(f'"{sys.executable}" -m pip install rich')
    from rich.console import Console
    from rich import print
    from rich.table import Table

console = Console()

# Константы путей (следуем принятой структуре проекта)
BASE_DIR = r"F:\Python Projets\Report"
DATA_WORK = BASE_DIR + r"\Data_work"
DATA_BACKUP = BASE_DIR + r"\Data_Backup"
NAME_JSON = DATA_WORK + r"\name_clients.json"

# Пути к справочникам
REF_STOCKS_XLSX = r"F:\Python Projets\Report\dictionaries\reference_stocks\reference_stocks_etf.xlsx"
REF_BONDS_XLSX  = r"F:\Python Projets\Report\dictionaries\reference_bonds\reference_bonds.xlsx"
REF_SP_XLSX     = r"F:\Python Projets\Report\dictionaries\reference_structured\TS\TS.xlsx"
REF_SP_PDF_DIR  = r"F:\Python Projets\Report\dictionaries\reference_structured\TS"

# ---------- Утилиты ----------

def parse_payload_name_from_filename(path: Path) -> Tuple[str, str, str]:
    """
    Из имени файла вида:
      isin_Фамилия И.О._DD.MM.YYYY__DD.MM.YYYY.json
    извлекает (client_str, start_date, end_date).
    Возвращает строки без изменений.
    """
    name = path.name
    # Жёсткое соответствие шаблону
    m = re.fullmatch(r"isin_(.+)_(\d{2}\.\d{2}\.\d{4})__(\d{2}\.\d{2}\.\d{4})\.json", name)
    if not m:
        raise ValueError(f"Имя файла не соответствует шаблону: {name}")
    client_str, start_date, end_date = m.group(1), m.group(2), m.group(3)
    return client_str, start_date, end_date


def load_client_isins(path: Path) -> Tuple[str, Dict[str, str], List[str]]:
    """
    Читает входной JSON и возвращает:
      client: строка из JSON (как есть)
      period: словарь {"start_date": "...", "end_date": "..."}
      isin_list: список ISIN
    Бросает осмысленные ошибки при проблемах с чтением/структурой.
    """
    try:
        with path.open("r", encoding="utf-8") as f:
            data = json.load(f)
    except FileNotFoundError:
        raise FileNotFoundError(f"Файл не найден: {path}")
    except json.JSONDecodeError as e:
        raise ValueError(f"Ошибка парсинга JSON в {path}: {e}")

    if not isinstance(data, dict):
        raise ValueError("Ожидался объект JSON верхнего уровня")

    client = data.get("client")
    period = data.get("period")
    isins = data.get("isin")

    if not isinstance(client, str) or not client.strip():
        raise ValueError("Поле 'client' отсутствует или пустое")

    if not isinstance(period, dict) or "start_date" not in period or "end_date" not in period:
        raise ValueError("Поле 'period' отсутствует или неполное")

    if not isinstance(isins, list) or not all(isinstance(x, str) for x in isins):
        raise ValueError("Поле 'isin' должно быть списком строк")

    return client, {"start_date": period["start_date"], "end_date": period["end_date"]}, isins


def find_input_payload(data_work: str) -> Path:
    """
    Ищет входной файл по маске isin_*.json в DATA_WORK.
    Если в папке есть файлы других клиентов, они перемещаются в Data_Backup.
    Для текущего клиента допускается ровно один файл; иначе — ошибка.
    При отсутствии name_clients.json сохраняется прежняя логика выбора.
    """
    pattern = os.path.join(data_work, "isin_*.json")
    files = [Path(p) for p in glob(pattern) if os.path.isfile(p)]

    if not files:
        console.print(f"[red]❌ Во входной папке [/red][bright_cyan]{data_work}[/bright_cyan][red] нет файлов по маске isin_*.json[/red]")
        sys.exit(1)

    # Пытаемся определить текущего клиента
    client = _read_current_client_from_namejson()

    # Если клиента определить не удалось — работаем по прежней схеме (выбрать один, остальные в резерв)
    if not client:
        keep, to_archive = _pick_isin_to_keep(files)
        console.print(f"[yellow]⚠️ Не удалось определить клиента по name_clients.json. "
                      f"Оставляю самый подходящий:[/yellow] [bright_cyan]{keep.name}[/bright_cyan]")
        _ensure_dir(Path(DATA_BACKUP))
        for old in to_archive:
            moved = _archive_path_to_backup(old)
            console.print(f"[yellow]↳ Перемещён в резерв:[/yellow] [bright_cyan]{moved}[/bright_cyan]")
        return keep

    # Фильтрация по текущему клиенту
    matching = [p for p in files if f"isin_{client}_" in p.name]
    foreign  = [p for p in files if p not in matching]

    # Все "чужие" входные JSON — в резерв
    if foreign:
        _ensure_dir(Path(DATA_BACKUP))
        for old in foreign:
            moved = _archive_path_to_backup(old)
            console.print(f"[yellow]⚠️ Найден входной JSON другого клиента, перемещён в резерв:[/yellow] [bright_cyan]{moved}[/bright_cyan]")

    # Проверки по текущему клиенту
    if not matching:
        console.print(f"[red]❌ Для клиента [/red][bright_cyan]{client}[/bright_cyan][red] не найдено ни одного файла "
                      f"isin_{client}_*.json в [/red][bright_cyan]{data_work}[/bright_cyan][red].[/red]")
        sys.exit(1)

    if len(matching) > 1:
        console.print(f"[red]❌ Найдено несколько входных файлов для клиента [/red][bright_cyan]{client}[/bright_cyan][red]:[/red]")
        for p in matching:
            console.print(f"[bright_cyan]  - {p.name}[/bright_cyan]")
        console.print("[yellow]Удали лишние или перемести их в Data_Backup и повтори запуск.[/yellow]")
        sys.exit(1)

    # Ровно один корректный файл
    return matching[0]

# === Автоустановка openpyxl ===
try:
    from openpyxl import load_workbook
except ImportError:
    os.system(f'"{sys.executable}" -m pip install openpyxl')
    from openpyxl import load_workbook


def _norm_isin(s: str) -> str:
    return (s or "").strip().upper()


def load_reference_stocks(xlsx_path: str) -> dict:
    """
    Лист: 'акции_etf'
    Колонки: A=ISIN, B=Тикер, C=Название, D=Тип
    Возврат: { ISIN: {"ticker": str, "type": str, "name": str} }
    """
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    if "акции_etф" in wb.sheetnames:
        ws = wb["акции_etф"]
    else:
        ws = wb["акции_etf"]  # на случай опечаток регистра/раскладки
    ref = {}
    # пропускаем заголовок (первая строка)
    for row in ws.iter_rows(min_row=2, values_only=True):
        isin, ticker, name, typ = row[:4]
        isin = _norm_isin(isin)
        if not isin:
            continue
        ref[isin] = {
            "ticker": (ticker or "").strip(),
            "type": (typ or "").strip(),
            "name": (name or "").strip(),
        }
    return ref


def load_reference_bonds(xlsx_path: str) -> dict:
    """
    Лист: 'bonds'
    Колонки: A=ISIN, B=Название инструмента
    Возврат: { ISIN: {"name": str} }
    """
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb["bonds"]
    ref = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        isin, name = row[:2]
        isin = _norm_isin(isin)
        if not isin:
            continue
        ref[isin] = {"name": (name or "").strip()}
    return ref


def load_reference_structured(xlsx_path: str, pdf_dir: str) -> dict:
    """
    Лист: 'TS'
    Колонки: B=ISIN, C=ссылка (необязательна для нас)
    Возврат: { ISIN: {"pdf_path": <str|None>} }
    PDF располагаются в pdf_dir и именуются '<ISIN>.pdf'
    """
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb["TS"]
    ref = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        # row: (N, ISIN, LINK, ...)
        _, isin, _ = (row + (None,))[:3]
        isin = _norm_isin(isin)
        if not isin:
            continue
        pdf_path = os.path.join(pdf_dir, f"{isin}.pdf")
        ref[isin] = {"pdf_path": pdf_path if os.path.isfile(pdf_path) else None}
    return ref


def match_isins(
    isins: List[str],
    ref_stocks: dict,
    ref_bonds: dict,
    ref_struct: dict,
) -> Tuple[List[dict], List[dict], List[dict], List[str]]:
    """
    Сопоставляет ISIN из входного JSON со справочниками.
    Приоритет строго такой (один ISIN → максимум в одну группу):
      1) Акции/ETF
      2) Облигации
      3) Структурные продукты
      4) Иначе — в 'misses'
    Возвращает кортеж списков:
      hits_stocks: list[{"isin","ticker","type"}]
      hits_bonds:  list[{"isin","name"}]
      hits_sp:     list[{"isin","type","pdf_path"}]  # type всегда "СТРУКТУРНЫЙ ПРОДУКТ"
      misses:      list[isin]
    """
    seen = set()
    hits_stocks: List[dict] = []
    hits_bonds: List[dict] = []
    hits_sp: List[dict] = []
    misses: List[str] = []

    for raw in isins:
        isin = (raw or "").strip().upper()
        if not isin or isin in seen:
            continue
        seen.add(isin)

        # 1) Stocks/ETF
        s = ref_stocks.get(isin)
        if s:
            hits_stocks.append({
                "isin": isin,
                "ticker": s.get("ticker", ""),
                "name": s.get("name", ""),
                "type": s.get("type", ""),
            })
            continue

        # 2) Bonds
        b = ref_bonds.get(isin)
        if b:
            hits_bonds.append({
                "isin": isin,
                "name": b.get("name", ""),
            })
            continue

        # 3) Structured products
        sp = ref_struct.get(isin)
        if sp:
            hits_sp.append({
                "isin": isin,
                "type": "СТРУКТУРНЫЙ ПРОДУКТ",
                "pdf_path": sp.get("pdf_path"),
            })
            continue

        # 4) Не найден ни в одном справочнике
        misses.append(isin)

    return hits_stocks, hits_bonds, hits_sp, misses

# ---------- Вспомогательные функции для Этапа 4 ----------

def _ts_suffix() -> str:
    return datetime.now().strftime("%Y%m%d_%H%M%S")

def _ensure_dir(p: Path) -> None:
    p.mkdir(parents=True, exist_ok=True)

def build_output_paths(client: str, period: dict) -> dict:
    """
    Возвращает словарь с именами выходных файлов и каталогом для SP.
    Все имена строго по шаблонам.
    """
    start = period["start_date"]
    end = period["end_date"]
    base = Path(DATA_WORK)
    return {
        "stocks_json": base / f"stock_etf_{client}_{start}__{end}.json",
        "bonds_json":  base / f"bonds_{client}_{start}__{end}.json",
        "sp_json":     base / f"sp_{client}_{start}__{end}.json",
        "noname_json": base / f"noname_isin_{client}_{start}__{end}.json",
        "sp_dir":      base / f"sp_{client}_{start}__{end}",
    }

def _archive_path_to_backup(path: Path) -> Path:
    """
    Перемещает файл/папку в DATA_BACKUP с суффиксом '_резерв_YYYYMMDD_HHMMSS'.
    Возвращает путь в бэкапе.
    """
    _ensure_dir(Path(DATA_BACKUP))
    suffix = _ts_suffix()
    target_name = f"{path.stem}_резерв_{suffix}{path.suffix}" if path.is_file() else f"{path.name}_резерв_{suffix}"
    target = Path(DATA_BACKUP) / target_name
    shutil.move(str(path), str(target))
    return target

def archive_existing_outputs(paths: dict) -> None:
    """
    Если выходные JSON уже существуют — переместить в Data_Backup.
    Если папка SP уже существует — также переместить в Data_Backup.
    """
    # JSON-файлы
    for key in ("stocks_json", "bonds_json", "sp_json", "noname_json"):
        p = Path(paths[key])
        if p.exists():
            moved = _archive_path_to_backup(p)
            console.print(f"[yellow]⚠️ Найден существующий файл, перемещен в резерв:[/yellow] [bright_cyan]{moved}[/bright_cyan]")

    # Папка SP
    sp_dir = Path(paths["sp_dir"])
    if sp_dir.exists():
        moved = _archive_path_to_backup(sp_dir)
        console.print(f"[yellow]⚠️ Найдена существующая папка TermSheets, перемещена в резерв:[/yellow] [bright_cyan]{moved}[/bright_cyan]")

def find_previous_sp_dirs(client: str, keep_dir: Path) -> list[Path]:
    """
    Ищет в Data_work все каталоги вида 'sp_{client}_*',
    КРОМЕ каталога с именем keep_dir.name (целевой на текущий запуск).
    """
    pattern = Path(DATA_WORK) / f"sp_{client}_*"
    candidates = [Path(p) for p in glob(str(pattern))]
    return [p for p in candidates if p.is_dir() and p.name != keep_dir.name]

def find_all_sp_dirs_except(keep_dir: Path) -> list[Path]:
    """
    Находит в Data_work все каталоги по маске 'sp_*',
    КРОМЕ каталога с именем keep_dir.name (целевой текущего запуска).
    Используется для очистки Data_work от папок других клиентов/периодов.
    """
    pattern = Path(DATA_WORK) / "sp_*"
    candidates = [Path(p) for p in glob(str(pattern))]
    return [p for p in candidates if p.is_dir() and p.name != keep_dir.name]

def archive_dirs_to_backup(dirs: list[Path]) -> None:
    """
    Перемещает перечисленные каталоги в Data_Backup
    с суффиксом '_резерв_{YYYYMMDD_%H%M%S}'.
    """
    if not dirs:
        return
    _ensure_dir(Path(DATA_BACKUP))
    for src in dirs:
        moved = _archive_path_to_backup(src)
        console.print(f"[yellow]⚠️ Найдена старая папка TermSheets, перемещена в резерв:[/yellow] [bright_cyan]{moved}[/bright_cyan]")

def find_previous_jsons_for_client(client: str, period: dict, keep_paths: dict) -> list[Path]:
    """
    Возвращает список ВСЕХ JSON в Data_work для данного клиента по маскам:
      stock_etf_{client}_*.json
      bonds_{client}_*.json
      sp_{client}_*.json
      noname_isin_{client}_*.json
    КРОМЕ текущих целевых файлов из keep_paths (их имена исключаем).
    """
    base = Path(DATA_WORK)
    patterns = [
        base / f"stock_etf_{client}_*.json",
        base / f"bonds_{client}_*.json",
        base / f"sp_{client}_*.json",
        base / f"noname_isin_{client}_*.json",
    ]
    keep_names = {
        Path(keep_paths["stocks_json"]).name,
        Path(keep_paths["bonds_json"]).name,
        Path(keep_paths["sp_json"]).name,
        Path(keep_paths["noname_json"]).name,
    }
    results = []
    for pat in patterns:
        for p in glob(str(pat)):
            pth = Path(p)
            if pth.is_file() and pth.name not in keep_names:
                results.append(pth)
    return results

def archive_jsons_to_backup(files: list[Path]) -> None:
    """
    Перемещает перечисленные JSON-файлы в Data_Backup с суффиксом '_резерв_YYYYMMDD_HHMMSS'.
    """
    if not files:
        return
    _ensure_dir(Path(DATA_BACKUP))
    for src in files:
        moved = _archive_path_to_backup(src)
        console.print(f"[yellow]⚠️ Найден старый JSON, перемещен в резерв:[/yellow] [bright_cyan]{moved}[/bright_cyan]")

def find_foreign_jsons(client: str, keep_paths: dict) -> list[Path]:
    """
    Находит ВСЕ JSON-файлы в Data_work (stock_etf_*, bonds_*, sp_*, noname_isin_*),
    относящиеся к другим клиентам (имя файла НЕ содержит client),
    и не совпадающие с текущими целевыми путями из keep_paths.
    """
    base = Path(DATA_WORK)
    patterns = [
        base / "stock_etf_*.json",
        base / "bonds_*.json",
        base / "sp_*.json",
        base / "noname_isin_*.json",
    ]
    keep_names = {
        Path(keep_paths["stocks_json"]).name,
        Path(keep_paths["bonds_json"]).name,
        Path(keep_paths["sp_json"]).name,
        Path(keep_paths["noname_json"]).name,
    }
    results = []
    for pat in patterns:
        for p in glob(str(pat)):
            pth = Path(p)
            # чужой клиент и не текущие целевые имена
            if pth.is_file() and (client not in pth.name) and (pth.name not in keep_names):
                results.append(pth)
    return results


def find_foreign_sp_dirs(client: str, keep_dir: Path) -> list[Path]:
    """
    Находит все каталоги sp_* в Data_work, которые НЕ относятся к client
    и не совпадают с keep_dir (целевой каталог текущего запуска).
    """
    pattern = Path(DATA_WORK) / "sp_*"
    candidates = [Path(p) for p in glob(str(pattern))]
    return [p for p in candidates if p.is_dir() and (client not in p.name) and (p.name != keep_dir.name)]

def write_json_with_header(out_path: Path, client: str, period: dict, items: list) -> None:
    payload = {
        "client": client,
        "period": {"start_date": period["start_date"], "end_date": period["end_date"]},
        "items": items,
    }
    with out_path.open("w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)
    console.print(f"[green]📝 JSON записан:[/green] [bright_cyan]{out_path}[/bright_cyan]")

def copy_termsheets(hits_sp: list[dict], target_dir: Path) -> tuple[int, int]:
    """
    Копирует существующие PDF по именам ISIN в целевой каталог.
    Возвращает (скопировано, отсутствуют).
    """
    _ensure_dir(target_dir)
    copied = 0
    missing = 0
    for rec in hits_sp:
        pdf = rec.get("pdf_path")
        isin = rec.get("isin", "")
        if pdf and os.path.isfile(pdf):
            dst = target_dir / f"{isin}.pdf"
            shutil.copy2(pdf, dst)
            copied += 1
        else:
            console.print(f"[yellow]⚠️ TermSheet не найден для ISIN:[/yellow] [bright_cyan]{isin}[/bright_cyan]")
            missing += 1
    return copied, missing

def _read_current_client_from_namejson() -> str | None:
    """
    Возвращает client_name из Data_work/name_clients.json, либо None.
    """
    try:
        with open(NAME_JSON, "r", encoding="utf-8") as f:
            data = json.load(f)
        client = (data.get("client_name") or "").strip()
        return client or None
    except Exception:
        return None

def _pick_isin_to_keep(files: list[Path]) -> tuple[Path, list[Path]]:
    """
    Из списка isin_*.json выбирает один, который оставляем, и список остальных для архивации.
    Приоритет:
      1) если есть name_clients.json — берем самый свежий файл, где имя содержит 'isin_{client}_'
      2) иначе — самый свежий по времени изменения (mtime)
    """
    assert files, "files must be non-empty"
    client = _read_current_client_from_namejson()

    keep: Path | None = None
    if client:
        matching = [p for p in files if f"isin_{client}_" in p.name]
        if matching:
            keep = max(matching, key=lambda p: p.stat().st_mtime)

    if keep is None:
        keep = max(files, key=lambda p: p.stat().st_mtime)

    to_archive = [p for p in files if p != keep]
    return keep, to_archive

# ---------- Точка входа ----------

def main(argv: Optional[List[str]] = None) -> int:
    try:
        console.print("[bold green]🧭 map_instruments — Этап 1 (каркас)[/bold green]")
        console.print(f"[bright_cyan]Поиск входного файла в: {DATA_WORK}[/bright_cyan]")

        input_path = find_input_payload(DATA_WORK)

        # Имя файла → ожидаемые client/start/end (по имени)
        client_from_name, start_from_name, end_from_name = parse_payload_name_from_filename(input_path)

        console.print(f"[green]✅ Найден входной JSON: [/green][bright_cyan]{input_path.name}[/bright_cyan]")
        console.print(f"[green]↳ Ожидается из имени: client=[/green][bright_cyan]{client_from_name}[/bright_cyan][green], "
                      f"period=[/green][bright_cyan]{start_from_name}..{end_from_name}[/bright_cyan]")

        # Фактическое содержимое JSON
        client, period, isins = load_client_isins(input_path)
        console.print(f"[green]✅ Загружен JSON. Клиент:[/green] [bright_cyan]{client}[/bright_cyan]")
        console.print(f"[green]↳ Период:[/green] [bright_cyan]{period['start_date']}..{period['end_date']}[/bright_cyan]")
        console.print(f"[green]↳ Кол-во ISIN:[/green] [bright_cyan]{len(isins)}[/bright_cyan]")

        # Загрузка справочников (только чтение, без сопоставления)
        console.print(f"[green]🔄 Загрузка справочников…[/green]")
        console.print(f"[green]↳ Stocks/ETF:[/green] [bright_cyan]{REF_STOCKS_XLSX}[/bright_cyan]")
        stocks = load_reference_stocks(REF_STOCKS_XLSX)
        console.print(f"[green]   Загружено записей:[/green] [bright_cyan]{len(stocks)}[/bright_cyan]")

        console.print(f"[green]↳ Bonds:[/green] [bright_cyan]{REF_BONDS_XLSX}[/bright_cyan]")
        bonds = load_reference_bonds(REF_BONDS_XLSX)
        console.print(f"[green]   Загружено записей:[/green] [bright_cyan]{len(bonds)}[/bright_cyan]")

        console.print(f"[green]↳ Structured (TS):[/green] [bright_cyan]{REF_SP_XLSX}[/bright_cyan]")
        structured = load_reference_structured(REF_SP_XLSX, REF_SP_PDF_DIR)
        console.print(f"[green]   Загружено записей:[/green] [bright_cyan]{len(structured)}[/bright_cyan]")

        # Сопоставление ISIN по справочникам (без записи на диск)
        hits_stocks, hits_bonds, hits_sp, misses = match_isins(isins, stocks, bonds, structured)

        console.print("[green]🧩 Результат сопоставления:[/green]")
        console.print(f"  Акции/ETF: [bright_cyan]{len(hits_stocks)}[/bright_cyan]")
        console.print(f"  Облигации: [bright_cyan]{len(hits_bonds)}[/bright_cyan]")
        console.print(f"  Структурные продукты: [bright_cyan]{len(hits_sp)}[/bright_cyan]")
        console.print(f"  Неизвестные (noname): [bright_cyan]{len(misses)}[/bright_cyan]")

        # Постоянный предпросмотр результатов (по 3 категориям). Записи не пишутся на диск.
        preview_limit = 20

        def _render_table(title: str, columns: list[str], rows: list[list[str]]):
            if not rows:
                console.print(f"[yellow]{title}: нет записей[/yellow]")
                return
            table = Table(title=title, show_lines=False)
            for col in columns:
                # номер колонки и короткие поля делаем no_wrap для аккуратного вида
                if col in ("№", "ISIN", "Ticker", "Type"):
                    table.add_column(col, no_wrap=True)
                else:
                    table.add_column(col)
            for i, r in enumerate(rows[:preview_limit], 1):
                table.add_row(str(i), *r)
            console.print(table)

        # 1) Предпросмотр: Акции/ETF
        stock_rows = [
            [rec.get("isin", ""), rec.get("ticker", ""), rec.get("name", ""), rec.get("type", "")]
            for rec in hits_stocks
        ]
        _render_table("Предпросмотр stock_etf (будущий JSON)", ["№", "ISIN", "Ticker", "Name", "Type"], stock_rows)

        # 2) Предпросмотр: Облигации
        bond_rows = [
            [rec.get("isin", ""), rec.get("name", "")]
            for rec in hits_bonds
        ]
        _render_table("Предпросмотр bonds (будущий JSON)", ["№", "ISIN", "Name"], bond_rows)

        # 3) Предпросмотр: Структурные продукты
        sp_rows = [
            [rec.get("isin", ""), rec.get("type", "СТРУКТУРНЫЙ ПРОДУКТ")]
            for rec in hits_sp
        ]
        _render_table("Предпросмотр structured (будущий JSON)", ["№", "ISIN", "Type"], sp_rows)

        # === Этап 4: запись выходных JSON и копирование TermSheets ===
        console.print("[green]💾 Формирование выходов (JSON + TermSheets)…[/green]")

        # Построить пути и имена
        paths = build_output_paths(client, period)

        # Архивировать прошлые результаты (если есть)
        archive_existing_outputs(paths)

        # Архивирование всех прошлых JSON этого клиента (с другими периодами), кроме текущих
        old_jsons = find_previous_jsons_for_client(client, period, paths)
        archive_jsons_to_backup(old_jsons)

        # Архивировать все прошлые папки TermSheets этого клиента (с другими периодами), кроме текущей
        old_sp_dirs = find_previous_sp_dirs(client, paths["sp_dir"])
        archive_dirs_to_backup(old_sp_dirs)

        # Архивировать JSON и папки других клиентов (полная уборка чужих артефактов)
        foreign_jsons = find_foreign_jsons(client, paths)
        archive_jsons_to_backup(foreign_jsons)

        foreign_sp_dirs = find_foreign_sp_dirs(client, paths["sp_dir"])
        archive_dirs_to_backup(foreign_sp_dirs)

        # Запись трех основных JSON
        write_json_with_header(paths["stocks_json"], client, period, hits_stocks)
        write_json_with_header(paths["bonds_json"],  client, period, hits_bonds)
        write_json_with_header(paths["sp_json"],     client, period, hits_sp)

        # Запись noname JSON при наличии пропусков
        if misses:
            write_json_with_header(paths["noname_json"], client, period, [{"isin": m} for m in misses])
        else:
            console.print("[green]✅ Неизвестных ISIN нет — noname JSON не создавался[/green]")

        # Копирование TermSheets
        copied, missing = copy_termsheets(hits_sp, paths["sp_dir"])
        console.print(f"[green]📦 Папка TermSheets:[/green] [bright_cyan]{paths['sp_dir']}[/bright_cyan]")
        console.print(f"[green]↳ Скопировано PDF:[/green] [bright_cyan]{copied}[/bright_cyan]; [yellow]Отсутствуют:[/yellow] [bright_cyan]{missing}[/bright_cyan]")

        console.print("[yellow]Этап 4 завершён: выходные JSON созданы, TermSheets скопированы (старые результаты отправлены в Data_Backup).[/yellow]")
        return 0

    except KeyboardInterrupt:
        console.print("\n[red]Операция прервана пользователем[/red]")
        return 1
    except Exception as e:
        console.print(f"[red]❌ Критическая ошибка: {e}[/red]")
        return 1


if __name__ == "__main__":
    sys.exit(main())
