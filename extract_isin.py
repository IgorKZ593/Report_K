#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Модуль извлечения уникальных валидных ISIN из Excel-отчетов.
Извлекает ISIN из листа 'портфель', валидирует по ISO 6166, формирует JSON.
"""

import os
import sys
import json
from glob import glob
import argparse
from pathlib import Path
from typing import Optional, List, Tuple
from datetime import datetime

# === Автоустановка rich (в первую очередь) ===
try:
    from rich.console import Console
    from rich import print
except ImportError:
    print("Устанавливаю rich для цветного вывода...")
    os.system(f'"{sys.executable}" -m pip install rich')
    from rich.console import Console
    from rich import print

# === Автоустановка openpyxl ===
try:
    from openpyxl import load_workbook
except ImportError:
    print("[bold yellow]Устанавливаю openpyxl...[/bold yellow]")
    os.system(f'"{sys.executable}" -m pip install openpyxl')
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("[bold red]Модуль openpyxl не установлен. Установите вручную: pip install openpyxl[/bold red]")
        sys.exit(1)

# Константы путей
BASE_DIR = r"F:\Python Projets\Report"
DATA_IN = BASE_DIR + r"\Data_in"
DATA_WORK = BASE_DIR + r"\Data_work"
DATA_BACKUP = BASE_DIR + r"\Data_Backup"
NAME_JSON = DATA_WORK + r"\name_clients.json"
DATES_JSON = DATA_WORK + r"\report_dates.json"

# Инициализация rich console
console = Console()


def ensure_dependencies() -> None:
    """Гарантирует наличие rich и openpyxl; при отсутствии — устанавливает через pip."""
    # Зависимости уже проверены в импортах выше
    pass


def load_json(path: str) -> dict:
    """Загружает JSON c UTF-8 и возвращает dict; на ошибке — понятное исключение."""
    try:
        with open(path, 'r', encoding='utf-8') as f:
            return json.load(f)
    except FileNotFoundError:
        raise FileNotFoundError(f"Файл {path} не найден")
    except json.JSONDecodeError as e:
        raise ValueError(f"Ошибка парсинга JSON в {path}: {e}")
    except Exception as e:
        raise Exception(f"Ошибка чтения {path}: {e}")


def normalize_sheet_name(name: str) -> str:
    """Возвращает нормализованное имя листа: lower + strip + без двойных пробелов."""
    return ' '.join(name.strip().lower().split())


def find_input_workbook() -> Path:
    """Находит ровно один файл отчет_*.xlsx в DATA_IN.
    0 файлов — ошибка; >1 — перечислить и ошибка; иначе вернуть Path к файлу."""
    if not os.path.exists(DATA_IN):
        console.print(f"[red]❌ Папка [/red][bright_cyan]{DATA_IN}[/bright_cyan][red] не найдена[/red]")
        sys.exit(1)
    
    # Поиск файлов по маске отчет_*.xlsx (регистрозависимо)
    pattern = os.path.join(DATA_IN, "отчет_*.xlsx")
    report_files = glob(pattern, recursive=False)
    
    # Фильтруем временные файлы Excel
    report_files = [f for f in report_files if not os.path.basename(f).startswith('~$')]
    
    if not report_files:
        console.print(f"[red]❌ В папке [/red][bright_cyan]{DATA_IN}[/bright_cyan][red] не найдено файлов по маске 'отчет_*.xlsx'[/red]")
        sys.exit(1)
    
    if len(report_files) > 1:
        console.print(f"[red]❌ Папка [/red][bright_cyan]{DATA_IN}[/bright_cyan][red] содержит более одного потенциального источника данных:[/red]")
        for file in report_files:
            console.print(f"[bright_cyan]  - {os.path.basename(file)}[/bright_cyan]")
        console.print("[yellow]⚠️  Просьба удалить лишние файлы[/yellow]")
        sys.exit(1)
    
    return Path(report_files[0])


def open_workbook(ws_path: Path):
    """Открывает книгу openpyxl (read-only=False) и возвращает объект workbook."""
    try:
        return load_workbook(ws_path, read_only=False)
    except Exception as e:
        console.print(f"[red]❌ Ошибка открытия файла [/red][bright_cyan]{ws_path.name}[/bright_cyan][red]: {e}[/red]")
        sys.exit(1)


def find_portfolio_sheet(wb):
    """Возвращает лист 'портфель' с нечувствительностью к регистру/пробелам; иначе ошибка."""
    sheet_dict = {normalize_sheet_name(sheet.title): sheet for sheet in wb.worksheets}
    target_name = normalize_sheet_name("портфель")
    
    if target_name not in sheet_dict:
        console.print(f"[red]❌ Лист 'портфель' не найден[/red]")
        console.print(f"[bright_cyan]Доступные листы: {[s.title for s in wb.worksheets]}[/bright_cyan]")
        sys.exit(1)
    
    return sheet_dict[target_name]


def find_isin_column(ws) -> int:
    """Находит индекс столбца по заголовку 'ISIN' в 1-й строке (casefold+strip); иначе ошибка."""
    for col_idx, cell in enumerate(ws[1], start=1):
        if str(cell.value).strip().casefold() == "isin":
            return col_idx
    
    console.print("[red]❌ Столбец 'ISIN' не найден в первой строке[/red]")
    headers = [(cell.value or "") for cell in ws[1]]
    console.print(f"[cyan]Заголовки: {headers}[/cyan]")
    sys.exit(1)


def luhn_check_isin(isin: str) -> bool:
    """Выполняет Luhn-проверку для ISIN (после замены букв на числа A=10..Z=35)."""
    # Преобразуем буквы в числа: A=10, B=11, ..., Z=35
    digits = ""
    for char in isin:
        if char.isalpha():
            digits += str(ord(char) - ord('A') + 10)
        else:
            digits += char
    
    # Luhn алгоритм: справа налево, удваиваем каждую вторую цифру
    total = 0
    for i, digit in enumerate(reversed(digits)):
        num = int(digit)
        if i % 2 == 1:  # Каждая вторая цифра справа
            num *= 2
            if num > 9:
                num = sum(int(d) for d in str(num))
        total += num
    
    return total % 10 == 0


def validate_isin(isin: str) -> bool:
    """Проверяет формат ISIN: длина 12, 2 буквы + 9 алфанум + контрольная цифра; затем Luhn."""
    if not isin or not isinstance(isin, str):
        return False
    
    isin = isin.strip().upper()
    
    # Проверка длины
    if len(isin) != 12:
        return False
    
    # Проверка формата: 2 буквы + 9 алфанум + 1 цифра
    if not (isin[:2].isalpha() and 
            isin[2:11].isalnum() and 
            isin[11].isdigit()):
        return False
    
    # Luhn проверка
    return luhn_check_isin(isin)


def read_isins(ws, col_idx: int) -> List[str]:
    """Считывает значения со 2-й строки до последней непустой, пропуская пустые; возвращает список строк."""
    isins = []
    for row in range(2, ws.max_row + 1):
        cell_value = ws.cell(row=row, column=col_idx).value
        if cell_value and str(cell_value).strip():
            isins.append(str(cell_value).strip())
    
    return isins


def unique_preserve_order(items: List[str]) -> Tuple[List[str], int]:
    """Возвращает (уникальный_список, число_выброшенных_дублей), сохраняя порядок первых вхождений."""
    seen = set()
    unique_items = []
    duplicates = 0
    
    for item in items:
        if item not in seen:
            seen.add(item)
            unique_items.append(item)
        else:
            duplicates += 1
    
    return unique_items, duplicates


def build_client_short(name_json: dict) -> Tuple[str, str]:
    """Из name_clients.client_name формирует:
       - client_for_json: исходную строку client_name без изменений (для JSON)
       - client_for_filename: 'Фамилия И.О.' (с пробелом между фамилией и инициалами)
       Поддерживаются оба варианта входа:
       1) 'Фамилия Имя Отчество' → 'Фамилия И.О.'
       2) 'Фамилия И.О.' → 'Фамилия И.О.' (без потери второй инициалы)
    """
    client_name = (name_json.get("client_name") or "").strip()
    if not client_name:
        raise ValueError("Поле 'client_name' не найдено в name_clients.json")

    parts = client_name.split(maxsplit=1)
    if len(parts) < 2:
        raise ValueError(f"Недостаточно данных для построения инициалов: {client_name}")

    surname, rest = parts[0], parts[1]

    # Извлекаем именно буквы из хвоста имени (работает и для 'Имя Отчество', и для 'И.О.')
    letters = [ch for ch in rest if ch.isalpha()]
    if len(letters) < 2:
        raise ValueError(f"Не удалось получить две инициалы из: {client_name}")

    initials = f"{letters[0].upper()}.{letters[1].upper()}."

    client_for_json = client_name                  # без изменений, как в name_clients.json
    client_for_filename = f"{surname} {initials}"  # 'Фамилия И.О.' (с пробелом)

    return client_for_json, client_for_filename


def build_output_filename(client_file: str, dates_json: dict) -> str:
    """Формирует имя файла JSON: isin_{client_file}_{start}__{end}.json (строго такой формат)."""
    start_date = dates_json.get("start_date", "")
    end_date = dates_json.get("end_date", "")
    
    if not start_date or not end_date:
        raise ValueError("Поля 'start_date' или 'end_date' не найдены в report_dates.json")
    
    return f"isin_{client_file}_{start_date}__{end_date}.json"


def find_previous_isin_jsons(client_file: str, keep_filename: str) -> list[Path]:
    """
    Ищет в Data_work все файлы вида 'isin_{client_file}_*.json', КРОМЕ точного имени keep_filename.
    Возвращает список путей (может быть пустым).
    """
    pattern = Path(DATA_WORK) / f"isin_{client_file}_*.json"
    candidates = [Path(p) for p in glob(str(pattern))]
    return [p for p in candidates if p.name != keep_filename]


def archive_files_to_backup(files: list[Path], yes: bool) -> None:
    """
    Перемещает перечисленные файлы в Data_Backup с суффиксом '_резерв_{YYYYMMDD_HHMMSS}'.
    Если yes=False, предварительно спрашивает один раз подтверждение на перемещение всех.
    При yes=True — перемещает молча.
    """
    if not files:
        return

    # Без вопросов, если --yes
    if yes:
        os.makedirs(DATA_BACKUP, exist_ok=True)
        for src in files:
            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            backup_name = f"{src.stem}_резерв_{ts}{src.suffix}"
            src.rename(Path(DATA_BACKUP) / backup_name)
            console.print(f"[bright_cyan]Перемещён: {src.name} → Data_Backup/{backup_name}[/bright_cyan]")
        return

    console.print("[yellow]Обнаружены предыдущие JSON-файлы isin для этого клиента:[/yellow]")
    for src in files:
        console.print(f"[bright_cyan]  - {src.name}[/bright_cyan]")

    while True:
        try:
            resp = input("Переместить их в Data_Backup? [Y/N]: ").strip().upper()
            if resp in ("Y", "YES"):
                os.makedirs(DATA_BACKUP, exist_ok=True)
                for src in files:
                    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
                    backup_name = f"{src.stem}_резерв_{ts}{src.suffix}"
                    src.rename(Path(DATA_BACKUP) / backup_name)
                    console.print(f"[bright_cyan]Перемещён: {src.name} → Data_Backup/{backup_name}[/bright_cyan]")
                break
            elif resp in ("N", "NO"):
                console.print("[grey]Оставили предыдущие файлы на месте[/grey]")
                break
            else:
                console.print("[yellow]Пожалуйста, введите Y или N[/yellow]")
        except (KeyboardInterrupt, EOFError):
            console.print("\n[red]Ввод прерван[/red]")
            sys.exit(1)


def handle_existing_output(path: Path, yes: bool) -> None:
    """Если файл существует:
       - yes=True: удалить;
       - yes=False: спросить Y/N в консоли; при N — перенести в Data_Backup с _резерв_YYYYMMDD_HHMMSS."""
    if not path.exists():
        return
    
    console.print(f"[yellow]⚠️  Файл уже существует:[/yellow] [bright_cyan]{path.name}[/bright_cyan]")
    
    if yes:
        path.unlink()
        console.print(f"[cyan]Файл удален (--yes режим)[/cyan]")
        return
    
    while True:
        try:
            response = input("Удалить существующий файл? [Y/N]: ").strip().upper()
            if response in ['Y', 'YES']:
                path.unlink()
                console.print(f"[cyan]Файл удален[/cyan]")
                return
            elif response in ['N', 'NO']:
                # Создаем Data_Backup если не существует
                os.makedirs(DATA_BACKUP, exist_ok=True)
                
                # Формируем имя резервного файла
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                backup_name = f"{path.stem}_резерв_{timestamp}{path.suffix}"
                backup_path = Path(DATA_BACKUP) / backup_name
                
                # Перемещаем файл
                path.rename(backup_path)
                console.print(f"[bright_cyan]Файл перенесен в Data_Backup: {backup_name}[/bright_cyan]")
                return
            else:
                console.print("[yellow]Пожалуйста, введите Y или N[/yellow]")
        except (KeyboardInterrupt, EOFError):
            console.print("\n[red]Ввод прерван[/red]")
            sys.exit(1)


def write_json(path: Path, payload: dict) -> None:
    """Записывает JSON c ensure_ascii=False и отступами."""
    try:
        with open(path, 'w', encoding='utf-8') as f:
            json.dump(payload, f, ensure_ascii=False, indent=2)
    except Exception as e:
        console.print(f"[red]❌ Ошибка записи JSON: {e}[/red]")
        sys.exit(1)


def main(argv: Optional[List[str]] = None) -> int:
    """Оркестратор: парсинг --yes, поиск книги, лист/столбец, чтение, валидация, уникализация, запись JSON."""
    try:
        # Парсинг аргументов
        parser = argparse.ArgumentParser(description="Извлечение ISIN из Excel-отчетов")
        parser.add_argument("--yes", "-y", action="store_true", 
                          help="Автоматически подтверждать все действия")
        args = parser.parse_args(argv)
        
        console.print("[bold green]🔍 Извлечение ISIN из Excel-отчета[/bold green]")
        
        # Шаг 1: Поиск входного файла
        console.print(f"[bright_cyan]Поиск файла отчета в: {DATA_IN}[/bright_cyan]")
        input_file = find_input_workbook()
        console.print(f"[green]✅ Найден файл: [/green][bright_cyan]{input_file.name}[/bright_cyan]")
        
        # Шаг 2: Открытие книги и поиск листа
        wb = open_workbook(input_file)
        portfolio_sheet = find_portfolio_sheet(wb)
        console.print(f"[green]✅ Найден лист: {portfolio_sheet.title}[/green]")
        
        # Шаг 3: Поиск столбца ISIN
        isin_col = find_isin_column(portfolio_sheet)
        console.print(f"[green]✅ Найден столбец ISIN (колонка {isin_col})[/green]")
        
        # Шаг 4: Чтение и валидация ISIN
        console.print("[cyan]Чтение и валидация ISIN...[/cyan]")
        raw_isins = read_isins(portfolio_sheet, isin_col)
        
        if not raw_isins:
            console.print("[red]❌ В столбце ISIN не найдено данных[/red]")
            return 1
        
        # Валидация ISIN
        valid_isins = []
        invalid_count = 0
        for isin in raw_isins:
            if validate_isin(isin):
                valid_isins.append(isin)
            else:
                invalid_count += 1
                console.print(f"[yellow]⚠️  Невалидный ISIN пропущен: {isin}[/yellow]")
        
        if not valid_isins:
            console.print("[red]❌ Валидных ISIN не найдено[/red]")
            return 1
        
        # Уникализация
        unique_isins, duplicates = unique_preserve_order(valid_isins)
        
        # Шаг 5: Загрузка метаданных
        try:
            name_data = load_json(NAME_JSON)
            dates_data = load_json(DATES_JSON)
        except Exception as e:
            console.print(f"[red]❌ Ошибка загрузки метаданных: {e}[/red]")
            return 1
        
        # Шаг 6: Формирование имени файла
        try:
            client_json, client_file = build_client_short(name_data)
            output_filename = build_output_filename(client_file, dates_data)
        except Exception as e:
            console.print(f"[red]❌ Ошибка формирования имени файла: {e}[/red]")
            return 1
        
        output_path = Path(DATA_WORK) / output_filename
        
        # Новый шаг: найти предыдущие JSON'ы для этого клиента и (опционально) переместить их в Data_Backup
        previous_jsons = find_previous_isin_jsons(client_file, output_filename)
        archive_files_to_backup(previous_jsons, args.yes)
        
        # Шаг 7: Обработка существующего файла
        if output_path.exists():
            handle_existing_output(output_path, args.yes)
        
        # Шаг 8: Создание папки Data_work если не существует
        os.makedirs(DATA_WORK, exist_ok=True)
        
        # Шаг 9: Формирование JSON
        payload = {
            "client": client_json,
            "period": {
                "start_date": dates_data["start_date"],
                "end_date": dates_data["end_date"]
            },
            "isin": unique_isins
        }
        
        # Шаг 10: Запись JSON
        write_json(output_path, payload)
        
        # Шаг 11: Вывод результатов
        console.print(f"\n[green]✅ Найдено валидных ISIN: {len(unique_isins)}[/green]")
        
        if duplicates > 0:
            console.print(f"[yellow]↺ Обнаружено и отброшено дублей: {duplicates}[/yellow]")
        
        if invalid_count > 0:
            console.print(f"[yellow]⚠️  Пропущено невалидных ISIN: {invalid_count}[/yellow]")
        
        console.print(f"[cyan]Список ISIN:[/cyan]")
        for i, isin in enumerate(unique_isins, 1):
            console.print(f"  {i:2d}. {isin}")
        
        console.print(f"\n[green]JSON сформирован:[/green] [bright_cyan]{output_path}[/bright_cyan]")
        
        return 0
        
    except KeyboardInterrupt:
        console.print("\n[red]❌ Операция прервана пользователем[/red]")
        return 1
    except Exception as e:
        console.print(f"[red]❌ Критическая ошибка: {e}[/red]")
        return 1


if __name__ == "__main__":
    sys.exit(main())
